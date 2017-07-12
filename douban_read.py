from bs4 import BeautifulSoup
import requests
import urllib.request
import openpyxl


headers = {'User-Agent':
               'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36'}

class Book:
    def __init__(self, title, author, pub, datetime, score, evalateNum, link):
        self.title = title
        self.author = author
        self.pub = pub
        self.datetime = datetime
        self.score = score
        self.evalateNum = evalateNum
        self.link = link

def getAllTags():
    url = 'https://book.douban.com/tag/'
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')
    tagsName = []
    for tagCol in soup.findAll('table', class_='tagCol'):
        for tag in tagCol.find_all('a'):
            tagsName.append(tag.text)
    return tagsName

def getOneTagBooks(tag):
    baseurl = 'https://book.douban.com/tag/'
    link = baseurl + urllib.request.quote(tag)
    books = []
    for page in range(2):
        url = link + '?start=' + str(page * 20) + '&type=T'
        print(url)
        r = requests.get(link, headers=headers)
        soup = BeautifulSoup(r.content, 'lxml')
        print("start craw page "+str(page))
        for bookinfo in soup.find_all("li", class_="subject-item"):
            if bookinfo == None:
                return books
            title = bookinfo.find('h2').a['title']
            pubinfo = bookinfo.find('div', class_='pub').text.strip().split('/')

            author = pubinfo[0].strip()
            datetime = pubinfo[-2].strip()
            pub = None
            for p in pubinfo:
                #print(p)
                if p.strip().endswith('出版'):
                    pub = p
            score = float(bookinfo.find('span', class_='rating_nums').text.strip())
            evalateNum = int(bookinfo.find('span', class_='pl').text.strip()[1:-4])
            booklink = bookinfo.find('h2').a['href']
            book = Book(title, author, pub, datetime, score, evalateNum, booklink)
            books.append(book)
            #print(book.title)
    return books

if __name__ == '__main__':

    #tagsName = getAllTags()
    books = getOneTagBooks('小说')
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '小说'
    sheet['A1'] = '书名'
    sheet['B1'] = '作者'
    sheet['C1'] = '出版社'
    sheet['D1'] = '评分'
    sheet['E1'] = '评价人数'
    sheet['F1'] = '出版日期'
    sheet['G1'] = '链接'
    i = 0
    print(len(books))
    for book in books:
        sheet.cell(row = i+2, column = 1).value = book.title
        sheet.cell(row = i+2, column = 2).value = book.author
        sheet.cell(row = i+2, column = 3).value = book.pub
        sheet.cell(row = i+2, column = 4).value = str(book.score)
        sheet.cell(row = i+2, column = 5).value = str(book.evalateNum)
        sheet.cell(row = i+2, column = 6).value = book.datetime
        sheet.cell(row = i+2, column = 7).value = book.link
        i+=1

    wb.save('douban_book.xlsx')
