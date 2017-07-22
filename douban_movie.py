from bs4 import BeautifulSoup
from urllib.request import urlopen
from urllib.request import Request
from urllib.error import *
import time
import openpyxl

headers = {'User-Agent':
               'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36'}


class DoubanMovieSpider:
    def __init__(self, url, movieurlfile):
        self.baseurl = url
        self.movieurlfile = movieurlfile
        self.movieinfos = []

    def getmovieurl(self):
        f = open(self.movieurlfile, 'w', encoding='utf-8')
        for page in range(10):
            url = self.baseurl + '?start=' + str(25 * (page)) + '&filter='
            try:
                res = Request(url, headers=headers)
                html = urlopen(url).read()
            except (HTTPError, URLError) as e:
                if hasattr(e, 'code'):
                    print(e.code)
                if hasattr(e, 'reason'):
                    print(e.reason)
            except Exception as e:
                print(e)
                continue

            soup = BeautifulSoup(html, 'lxml')

            for url in soup.find('ol', class_='grid_view').find_all('li'):
                movieurl = url.find('a')['href']
                # self.movieurls.append(movieurl)
                f.write(movieurl + '\n')
                # print(movieurl)
            time.sleep(1)
        f.close()

    def getMovieInfo(self):

        with open(self.movieurlfile, 'r', encoding='utf-8') as f:
            movieurls = f.readlines()
        for url in movieurls:
            print(url)
            movieinfo = []
            try:
                res = Request(url, headers=headers)
                html = urlopen(url).read().decode('utf-8')
            except (HTTPError, URLError) as e:
                if hasattr(e, 'code'):
                    print(e.code)
                if hasattr(e, 'reason'):
                    print(e.reason)
            except Exception as e:
                print(e)
                continue

            soup = BeautifulSoup(html, 'lxml')

            content = soup.find('div', {"id":"content"})

            title = content.h1.find('span').text.strip()
            movieinfo.append('片名: '+title)
            year = content.find('span', class_='year').text.strip()[1:-1]
            movieinfo.append('年份: '+year)
            info = content.find('div', {'id':'info'}).text.strip()
            #print(info.text.strip())
            ratingnum = content.find('div', {'class':'rating_self clearfix'}).strong.text.strip()
            #print(ratingnum)
            ratingsum = soup.find('span', {'property':'v:votes'}).text.strip()
            #print(ratingsum)
            movieinfo.append(info)
            movieinfo.append('评分: '+ratingnum)
            movieinfo.append('评价人数: '+ratingsum)
            self.movieinfos.append('\n'.join(movieinfo))
            #for t in self.movieinfos[0].split('\n'):
            #    print(t)
            time.sleep(5)

    def writetofile(self, filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '豆瓣电影top250'
        sheet['A1'] = '片名'
        sheet['B1'] = '年份'
        sheet['C1'] = '导演'
        sheet['D1'] = '编剧'
        sheet['E1'] = '主演'
        sheet['F1'] = '类型'
        sheet['G1'] = '制片国家/地区'
        sheet['H1'] = '语言'
        sheet['I1'] = '上映日期'
        sheet['J1'] = '片长'
        sheet['K1'] = '又名'
        sheet['L1'] = 'IMDB链接'
        sheet['M1'] = '评分'
        sheet['N1'] = '评价人数'

        i = 2
        for movie in self.movieinfos:
            movieinfo = movie[0].split('\n')
            #sheet.cell(row = i, column = 1).value = movieinfo[0]
            col = 1
            for t in movieinfo:
                sheet.cell(row = i, column = col).value = t.split(':')[1].strip()
                col += 1

        wb.save(filename)



douban = DoubanMovieSpider('https://movie.douban.com/top250', ' movieurl.txt')
douban.getMovieInfo()
douban.writetofile('douban_movie_top250.xlsx')
